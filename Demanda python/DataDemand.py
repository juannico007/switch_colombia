import csv
import datetime
import openpyxl
import pandas as pd
   
importedfile = openpyxl.load_workbook(filename = 'DemandaZonas.xlsx', read_only = True, keep_vba = False)
tabnames = importedfile.sheetnames 

##############################################################################
def is_empty(any_structure):
    if any_structure:
        return False
    else:
        return True
       
###########################################################################

read_file = pd.read_excel ('DemandaZonas.xlsx', sheet_name='DemandaMensual')
read_file.to_csv ('DemandaMensual.csv', index = None, header=True)

with open('DemandaMensual.csv') as csvfile:
    readCSV = csv.reader(csvfile, delimiter=',')
    dmdData = [[] for x in range(4)];  
    for row in readCSV:
        if is_empty(row) is True: break 
        for col in range(4):
            val = row[col]
            try: 
                val = float(row[col])
            except ValueError:
                pass
            dmdData[col].append(val)
for col in range(4):
    dmdData[col].pop(0)


# ###########################################################################

read_file = pd.read_excel ('DemandaZonas.xlsx', sheet_name='Reservas')
read_file.to_csv ('Reservas.csv', index = None, header=True)

with open('Reservas.csv') as csvfile:
    readCSV = csv.reader(csvfile, delimiter=',')
    dmdDataR = [];  
    for row in readCSV:
        if is_empty(row) is True: break 
        for col in range(1):
            val = row[col]
            try: 
                val = float(row[col])
            except ValueError:
                pass
            dmdDataR.append(val)
    
# ###########################################################################

read_file = pd.read_excel ('DemandaZonas.xlsx', sheet_name='Horario')
read_file = read_file.drop(read_file.columns[0], axis=1)
read_file.to_csv ('Horario.csv', index = None, header=True)


with open('Horario.csv') as csvfile:
    readCSV = csv.reader(csvfile, delimiter=',')
    horaData = [[] for x in range(16)];  
    for row in readCSV:
        if is_empty(row) is True: break 
        for col in range(16):
            val = row[col]
            try: 
                val = float(row[col])
            except ValueError:
                pass
            horaData[col].append(val)
# for col in range(columns-1):
#     horaData[col].pop(0) 

# ###########################################################################

read_file = pd.read_excel ('DemandaZonas.xlsx', sheet_name='Zonas')
read_file = read_file.drop(read_file.columns[0], axis=1)
read_file.to_csv ('Zonas.csv', index = None, header=True)

with open('Zonas.csv') as csvfile:
    readCSV = csv.reader(csvfile, delimiter=',')
    zonasData = [[] for x in range(15)];  
    for row in readCSV:
        if is_empty(row) is True: break 
        for col in range(15):
            val = row[col]
            try: 
                val = float(row[col])
            except ValueError:
                pass
            zonasData[col].append(val)
# for col in range(columns-1):
#     zonasData[col].pop(0) 

# ###########################################################################
month=[[] for x in range(12)]
for y in range(15):
    for m in range(12):
        month[m].append(dmdData[1][12*y+m])
    

# AR example
from statsmodels.tsa.arima.model import ARIMA
for m in range(12):
    for t in range(15):
        # contrived dataset
        data = month[m]
        # fit model
        model = ARIMA(data, order=(1, 1, 1))
        model_fit = model.fit()
        # make prediction
        yhat = model_fit.predict(len(data), len(data), typ='levels')
        month[m].append(yhat[0])

#Mese
q=[[sum(x)/3 for x in zip(month[0], month[1], month[2])],
   [sum(x)/3 for x in zip(month[3], month[10], month[11])],
   [sum(x)/3 for x in zip(month[7], month[8], month[9])],
   [sum(x)/3 for x in zip(month[4], month[5], month[6])]]

# Condiciones climáticas similares
# Q1	Ene, Feb, Mar
# Q2	Abr, Nov, Dic
# Q3	Ago, Sept, Oct
# Q4	May, Jun, Jul

promm=[[] for x in range(4)]
for qfila in range(4):  #4 cuartiles
    for y in range(6):  #6 bloques de 5 años
        val=[]
        for m in range(5): #Longitud de cada bloque
            val.append(q[qfila][5*y+m]) 
        promm[qfila].append(max(val)) #Max del prom por cuartil en esos 5 años

prommlist=[]
for y in range(6):      #6 bloques
    for qf in range(4): #por cuartil
        prommlist.append(promm[qf][y])

zondmq=[[] for x in range(15)]
for z in range(15):         # Recorre las 15 zonas
    for y in range(24):     # 24 cuartiles en total
        zondmq[z].append(zonasData[z][y+1]*prommlist[y])  #Desagregar por zonas

final=[[] for x in range(3)]
for z in range(15):         #15 zonas
    val=0
    for y in range(24):     #24 cuartiles
        for x in range(24): #24 horas
            final[0].append(horaData[z+1][0])
            final[1].append(val+1)
            val=val+1
            final[2].append(zondmq[z][y]*horaData[z+1][x+1])

from openpyxl import Workbook

# Create file
wb = Workbook()
dest_filename = 'colombia6/datos_python.xlsx'

ws0 = wb.active
ws0.title = "Summary"

ws0['A2'] = '24 horas';

###########################################################################

# write file results
ws1 = wb.create_sheet(title="Loads")
for k in range(len(final[0])):

    _ = ws1.cell(column=1, row=k+1, value=final[0][k])
    _ = ws1.cell(column=2, row=k+1, value=final[1][k])
    _ = ws1.cell(column=3, row=k+1, value=final[2][k])

###########################################################################

# write file results
ws1 = wb.create_sheet(title="Reservas")
for x in range(len(dmdDataR)):
    for k in range(576):

        _ = ws1.cell(column=1, row=(576*x)+k+1, value=dmdDataR[x])
        _ = ws1.cell(column=2, row=(576*x)+k+1, value=k+1)
        _ = ws1.cell(column=3, row=(576*x)+k+1, value=1)
    
# with open("loads.csv", "w") as f:
#     writer = csv.writer(f)
#     writer.writerows(final)
    
wb.save(filename = dest_filename)